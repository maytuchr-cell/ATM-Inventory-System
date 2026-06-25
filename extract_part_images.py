"""
Extract embedded part images from the 3 GRG catalogs, match each to a Part Number
by its anchored row, save to wwwroot/uploads/parts/, and update Part.ImagePath +
MainUnit + Remark in the SQLite DB.

Run with the API STOPPED (it writes directly to AtmInventory.db).
"""
import openpyxl, sqlite3, os, re, json

CATALOGS = [
    r'C:\Users\Admin\Downloads\Catalog D1-GRG Part_Update Oct 2023_Sub Unit Edition.xlsx',
    r'C:\Users\Admin\Downloads\Catalog D1-GRG Part_ICBC Project.xlsx',
    r'C:\Users\Admin\Downloads\Catalog D1-GRG Part_GSB ATM Project (1.7.2025).xlsx',
]
OUT_DIR = r'D:\ATMApi\ATM-Inventory-System\Backend\Api\wwwroot\uploads\parts'
DB_PATH = r'D:\ATMApi\ATM-Inventory-System\Backend\Api\AtmInventory.db'
SHEET   = 'Spare parts list'

# Column layout (1-based): C=Part Number(3), D=Description(4), E/F = Main/Sub Unit, I=Remark(9)
COL_PARTNO = 3

os.makedirs(OUT_DIR, exist_ok=True)

def safe(partno):
    return re.sub(r'[^A-Za-z0-9._-]', '_', str(partno).strip())

def detect_units(ws):
    """Return (main_col, sub_col, remark_col) by reading header row 2."""
    main_col = sub_col = remark_col = None
    for c in range(1, 15):
        v = ws.cell(row=2, column=c).value
        if not v: continue
        v = str(v).strip().lower()
        if v == 'main unit': main_col = c
        elif v == 'sub unit': sub_col = c
        elif v == 'remark': remark_col = c
    return main_col, sub_col, remark_col

# partNo -> {image: path, main: str, remark: str}
catalog_data = {}
img_count = 0

for f in CATALOGS:
    print(f'Processing: {os.path.basename(f)}')
    wb = openpyxl.load_workbook(f)  # not read_only — needed for images
    if SHEET not in wb.sheetnames:
        print('   no "Spare parts list" sheet, skip'); continue
    ws = wb[SHEET]
    main_col, sub_col, remark_col = detect_units(ws)

    # Build row -> partNo, and collect main/remark text per part
    row_partno = {}
    for r in range(3, ws.max_row + 1):
        pn = ws.cell(row=r, column=COL_PARTNO).value
        if pn and str(pn).strip():
            pn = str(pn).strip()
            row_partno[r] = pn
            entry = catalog_data.setdefault(pn, {})
            if main_col:
                mv = ws.cell(row=r, column=main_col).value
                if mv and str(mv).strip(): entry['main'] = str(mv).strip()
            if remark_col:
                rv = ws.cell(row=r, column=remark_col).value
                if rv and str(rv).strip(): entry['remark'] = str(rv).strip()

    # Extract images by anchor row
    imgs = getattr(ws, '_images', [])
    for img in imgs:
        try:
            anchor = img.anchor
            frm = getattr(anchor, '_from', None) or getattr(anchor, 'from', None)
            if frm is None: continue
            row0 = frm.row          # 0-based
            xl_row = row0 + 1       # 1-based
            # image may anchor a row or two below header offset — try exact, then +/-1
            pn = row_partno.get(xl_row) or row_partno.get(xl_row+1) or row_partno.get(xl_row-1)
            if not pn: continue
            data = img._data()
            ext = (img.format or 'jpeg').lower()
            if ext == 'jpeg': ext = 'jpg'
            fname = f'{safe(pn)}.{ext}'
            with open(os.path.join(OUT_DIR, fname), 'wb') as out:
                out.write(data)
            catalog_data.setdefault(pn, {})['image'] = f'/uploads/parts/{fname}'
            img_count += 1
        except Exception as e:
            pass
    wb.close()

print(f'Extracted {img_count} images for {len(catalog_data)} parts')

# ── Update DB ──
con = sqlite3.connect(DB_PATH)
cur = con.cursor()
updated_img = updated_meta = 0
for pn, d in catalog_data.items():
    sets, vals = [], []
    if d.get('image'):  sets.append('ImagePath = ?'); vals.append(d['image'])
    if d.get('main'):   sets.append('MainUnit = ?');  vals.append(d['main'])
    if d.get('remark'): sets.append('Remark = ?');    vals.append(d['remark'])
    if not sets: continue
    vals.append(pn)
    cur.execute(f'UPDATE Parts SET {", ".join(sets)} WHERE PartNo = ?', vals)
    if cur.rowcount > 0:
        if d.get('image'): updated_img += 1
        updated_meta += 1
con.commit()
# stats
total_with_img = cur.execute("SELECT COUNT(*) FROM Parts WHERE ImagePath IS NOT NULL AND ImagePath != ''").fetchone()[0]
con.close()
print(f'DB updated: {updated_img} parts got images, {updated_meta} parts got metadata')
print(f'Total parts with image now: {total_with_img}')
